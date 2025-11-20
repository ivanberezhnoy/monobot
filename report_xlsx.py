# report_xlsx.py

from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment


def write_xlsx(output_path: str, rows: List[Dict[str, Any]]) -> None:
    """
    rows — список словарей вида:
      {
        "_token_id": int,
        "_account_id": int,
        "token_name": str,
        "account_name": str,
        "datetime": str,
        "amount": float|str,
        "comment": str,
        "flow": "in"|"out",
        "account_flow_label": str,
      }

    Структура файла:

    [СМЕРЖЕННЫЙ ЗАГОЛОВОК ТОКЕНА (жирный, 14)]
    [СМЕРЖЕННЫЙ ЗАГОЛОВОК СЧЁТА]
    Дата и время | Сумма | Комментарий
    ... операции ...
    Итого по счёту ...
    ...
    Итого по токену ...
    """

    wb = Workbook()
    ws = wb.active

    DATE_COL = 1
    AMOUNT_COL = 2
    COMMENT_COL = 3
    LAST_COL = 3

    current_row = 1

    current_token_id = None
    current_token_name = ""
    token_total_in = 0.0
    token_total_out = 0.0

    current_account_id = None
    current_account_name = ""
    current_account_flow_label = ""
    account_total_in = 0.0
    account_total_out = 0.0

    def write_account_total():
        nonlocal current_row, account_total_in, account_total_out, current_account_name, current_account_id
        if current_account_id is None:
            return
        label = f"Итого по счёту {current_account_name}"
        wrote_any = False

        ws.cell(row=current_row, column=DATE_COL, value=f"{label} — входящие")
        ws.cell(row=current_row, column=AMOUNT_COL, value=round(account_total_in, 2))
        ws.cell(row=current_row, column=DATE_COL).font = Font(bold=True)
        ws.cell(row=current_row, column=AMOUNT_COL).font = Font(bold=True)
        wrote_any = True
        current_row += 1

        if account_total_out:
            ws.cell(row=current_row, column=DATE_COL, value=f"{label} — исходящие")
            ws.cell(row=current_row, column=AMOUNT_COL, value=round(account_total_out, 2))
            ws.cell(row=current_row, column=DATE_COL).font = Font(bold=True)
            ws.cell(row=current_row, column=AMOUNT_COL).font = Font(bold=True)
            current_row += 1
        elif not account_total_in:
            # если данных не было совсем, откатываем строку
            current_row -= 1
            wrote_any = False

        if wrote_any:
            current_row += 1  # пустая строка

        account_total_in = 0.0
        account_total_out = 0.0

    def write_token_total():
        nonlocal current_row, token_total_in, token_total_out, current_token_name, current_token_id
        if current_token_id is None:
            return
        label = f"Итого по токену {current_token_name}"

        ws.cell(row=current_row, column=DATE_COL, value=f"{label} — входящие")
        ws.cell(row=current_row, column=AMOUNT_COL, value=round(token_total_in, 2))
        ws.cell(row=current_row, column=DATE_COL).font = Font(bold=True, size=12)
        ws.cell(row=current_row, column=AMOUNT_COL).font = Font(bold=True, size=12)
        current_row += 1

        if token_total_out:
            ws.cell(row=current_row, column=DATE_COL, value=f"{label} — исходящие")
            ws.cell(row=current_row, column=AMOUNT_COL, value=round(token_total_out, 2))
            ws.cell(row=current_row, column=DATE_COL).font = Font(bold=True, size=12)
            ws.cell(row=current_row, column=AMOUNT_COL).font = Font(bold=True, size=12)
            current_row += 1

        current_row += 2  # пустые строки
        token_total_in = 0.0
        token_total_out = 0.0

    for row in rows:
        token_id = row["_token_id"]
        account_id = row["_account_id"]
        token_name = row["token_name"]
        account_name = row["account_name"]
        account_flow_label = row.get("account_flow_label", "")

        # смена токена
        if current_token_id is not None and token_id != current_token_id:
            write_account_total()
            write_token_total()
            current_account_id = None
            current_account_flow_label = ""

        if current_token_id != token_id:
            current_token_id = token_id
            current_token_name = token_name
            ws.merge_cells(
                start_row=current_row,
                start_column=DATE_COL,
                end_row=current_row,
                end_column=LAST_COL,
            )
            cell = ws.cell(row=current_row, column=DATE_COL, value=current_token_name)
            cell.font = Font(bold=True, size=14)
            cell.alignment = Alignment(horizontal="center")
            current_row += 2

        # смена счёта внутри токена
        if current_account_id is not None and account_id != current_account_id:
            write_account_total()
            current_account_flow_label = ""

        if current_account_id != account_id:
            current_account_id = account_id
            current_account_name = account_name
            current_account_flow_label = account_flow_label

            ws.merge_cells(
                start_row=current_row,
                start_column=DATE_COL,
                end_row=current_row,
                end_column=LAST_COL,
            )
            header_value = current_account_name
            if current_account_flow_label:
                header_value = f"{current_account_name} — {current_account_flow_label}"
            cell = ws.cell(row=current_row, column=DATE_COL, value=header_value)
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal="left")
            current_row += 1

            # заголовок таблицы
            ws.cell(row=current_row, column=DATE_COL, value="Дата и время")
            ws.cell(row=current_row, column=AMOUNT_COL, value="Сумма")
            ws.cell(row=current_row, column=COMMENT_COL, value="Комментарий")
            for col in range(DATE_COL, LAST_COL + 1):
                ws.cell(row=current_row, column=col).font = Font(bold=True)
            current_row += 1

        dt_str = row["datetime"]
        try:
            amt = float(row["amount"])
        except Exception:
            amt = float(str(row["amount"]).replace(",", "."))

        comment = row.get("comment", "")
        flow = row.get("flow", "in")

        ws.cell(row=current_row, column=DATE_COL, value=dt_str)
        amount_cell = ws.cell(row=current_row, column=AMOUNT_COL, value=amt)
        if flow == "out":
            amount_cell.font = Font(color="FFC00000")
        else:
            amount_cell.font = Font(color="FF008000")
        ws.cell(row=current_row, column=COMMENT_COL, value=comment)

        if flow == "out":
            account_total_out += amt
            token_total_out += amt
        else:
            account_total_in += amt
            token_total_in += amt

        current_row += 1

    # завершение последнего счёта/токена
    write_account_total()
    write_token_total()

    # автоширина
    for col_idx in range(1, LAST_COL + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = max_len + 2

    wb.save(output_path)
